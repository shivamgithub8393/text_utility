from flask import Flask, jsonify, request
from flask_cors import CORS, cross_origin
from openpyxl import load_workbook, Workbook
from datetime import datetime

# creating a Flask app
app = Flask(__name__)

cors = CORS(app, resources={r"/*": {"origins": "*"}})

# generated file name
output_file_name = 'output' # without extension

@cross_origin()
@app.route('/test', methods = ['GET', 'POST'])
def home():
    if(request.method == 'GET'):
        try:
            input_1 = request.args.get('input_1')
            input_2 = request.args.get('input_2')
            domain_selected = request.args.get('domain_selected')
            
            print("input_1: " + input_1, "input_2: " + input_2, " domain_selected : " + domain_selected)
            
            # append data into excel file
            # Specify the Workbook
            try:
                print("exists")
                wb_append = load_workbook("../output_files/"+ output_file_name+ '.xlsx')
            except:
                print("not exists")
                wb_append = Workbook()
            finally:
                pass
            
            sheet = wb_append.active
            row = (input_1, input_2, domain_selected, datetime.today())
            sheet.append(row)
            
            # Saving the data in our sample workbook/sheet
            wb_append.save('../output_files/'+ output_file_name + '.xlsx')
    
            return jsonify({'status': "success", 'data': "hello world"})
        except PermissionError:
            return jsonify({'status': "error", 'data': "Output file is open. First close it and try again"})
        except:
            return jsonify({'status': "error", 'data': "Something went wrong"})
  
# driver function
if __name__ == '__main__':
    app.run(debug = True)